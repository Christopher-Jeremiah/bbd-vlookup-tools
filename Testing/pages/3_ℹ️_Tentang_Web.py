import streamlit as st 
import pandas as pd 
import io 
import numpy as np 
from openpyxl.styles import PatternFill 

st.set_page_config(page_title="Data Automation Hub", layout="wide") 
st.title("⚙️ Hub Otomatisasi Data Universal") 

# ========================================== 
# 0. INISIALISASI MEMORI (BRANKAS) 
# ========================================== 
if 'id_kondisi' not in st.session_state: 
    st.session_state.id_kondisi = [1] 
if 'penghitung_id' not in st.session_state: 
    st.session_state.penghitung_id = 1 

if 'id_hapus' not in st.session_state: 
    st.session_state.id_hapus = [] 
if 'penghitung_id_hapus' not in st.session_state: 
    st.session_state.penghitung_id_hapus = 0 

if 'id_warna' not in st.session_state: 
    st.session_state.id_warna = [] 
if 'penghitung_id_warna' not in st.session_state: 
    st.session_state.penghitung_id_warna = 0 

if 'id_ganti' not in st.session_state: 
    st.session_state.id_ganti = [] 
if 'penghitung_id_ganti' not in st.session_state: 
    st.session_state.penghitung_id_ganti = 0 

if 'proses_selesai' not in st.session_state: 
    st.session_state.proses_selesai = False 
if 'df_file1' not in st.session_state: 
    st.session_state.df_file1 = None 
if 'df_file2' not in st.session_state: 
    st.session_state.df_file2 = None 

def tambah_kondisi(): 
    st.session_state.penghitung_id += 1 
    st.session_state.id_kondisi.append(st.session_state.penghitung_id) 

def hapus_kondisi(uid): 
    st.session_state.id_kondisi.remove(uid) 

def tambah_hapus(): 
    st.session_state.penghitung_id_hapus += 1 
    st.session_state.id_hapus.append(st.session_state.penghitung_id_hapus) 

def hapus_baris_hapus(uid): 
    st.session_state.id_hapus.remove(uid) 

def tambah_warna(): 
    st.session_state.penghitung_id_warna += 1 
    st.session_state.id_warna.append(st.session_state.penghitung_id_warna) 

def hapus_baris_warna(uid): 
    st.session_state.id_warna.remove(uid) 

def tambah_ganti(): 
    st.session_state.penghitung_id_ganti += 1 
    st.session_state.id_ganti.append(st.session_state.penghitung_id_ganti) 

def hapus_baris_ganti(uid): 
    st.session_state.id_ganti.remove(uid) 


# ========================================== 
# PINTU MASUK FILE 
# ========================================== 
col1, col2 = st.columns(2) 
with col1: 
    file_utama = st.file_uploader("1. Unggah Tabel Utama", type=["xlsx", "csv"]) 
    sheet_utama = None 
    header_utama = 0 
    xls_utama = None 
    
    if file_utama is not None: 
        if file_utama.name.endswith('.xlsx'): 
            xls_utama = pd.ExcelFile(file_utama) 
            sheet_utama = st.selectbox("Pilih Sheet (Utama):", xls_utama.sheet_names) 
        
        header_utama = st.number_input("Baris Header Utama (Mulai dari 0):", min_value=0, value=0, step=1) 

with col2: 
    file_ref = st.file_uploader("2. Unggah Tabel Kamus", type=["xlsx", "csv"]) 
    sheet_ref = None 
    header_ref = 0 
    xls_ref = None 
    
    if file_ref is not None: 
        if file_ref.name.endswith('.xlsx'): 
            xls_ref = pd.ExcelFile(file_ref) 
            sheet_ref = st.selectbox("Pilih Sheet (Kamus):", xls_ref.sheet_names) 
            
        header_ref = st.number_input("Baris Header Kamus (Mulai dari 0):", min_value=0, value=0, step=1) 

# ========================================== 
# ANTARMUKA DINAMIS (UI) 
# ========================================== 
if file_utama is not None and file_ref is not None: 
    
    try: 
        if file_utama.name.endswith('.csv'): 
            df_utama = pd.read_csv(file_utama, header=header_utama, dtype=str) 
        else: 
            df_utama = pd.read_excel(xls_utama, sheet_name=sheet_utama, header=header_utama, dtype=str) 
            
        if file_ref.name.endswith('.csv'): 
            df_ref = pd.read_csv(file_ref, header=header_ref, dtype=str) 
        else: 
            df_ref = pd.read_excel(xls_ref, sheet_name=sheet_ref, header=header_ref, dtype=str) 
    except Exception as e: 
        st.error(f"❌ Gagal membaca file. Pastikan letak baris Header yang Anda masukkan benar! Detail error: {e}") 
        st.stop() 

    def perbaiki_judul_tanggal(df): 
        kolom_rapi = [] 
        for col in df.columns: 
            col_str = str(col) 
            if ' 00:00:00' in col_str: 
                try: 
                    tgl_obj = pd.to_datetime(col_str) 
                    kolom_rapi.append(tgl_obj.strftime('%d-%b')) 
                except: 
                    kolom_rapi.append(col_str) 
            else: 
                kolom_rapi.append(col_str) 
        df.columns = kolom_rapi 
        return df 

    df_utama = perbaiki_judul_tanggal(df_utama) 
    df_ref = perbaiki_judul_tanggal(df_ref) 
    
    st.markdown("---") 
    
    # ------------------------------------------ 
    # UI TAHAP 1: PEMBERSIHAN DATA & GANTI NILAI 
    # ------------------------------------------ 
    st.subheader("🧹 Tahap 1: Pembersihan Awal & Ganti Data (Tabel Utama)") 
    
    daftar_kolom_utama = ["(Lewati)"] + list(df_utama.columns) 
    pilihan_hapus_kosong = st.selectbox("Hapus baris jika kolom ini kosong (Biasanya Baris Total):", daftar_kolom_utama) 
    
    st.markdown("---") 
    st.markdown("**Cari dan Ganti Nilai (Find & Replace Sebelum VLOOKUP):**") 
    
    for uid_g in st.session_state.id_ganti: 
        col_g1, col_g2, col_g3, col_g4, col_g5 = st.columns([2, 4, 2, 2, 1]) 
        
        with col_g1: 
            kol_terpilih = st.selectbox("Di kolom:", list(df_utama.columns), key=f"kolom_ganti_{uid_g}") 
            
        with col_g2: 
            nilai_unik = df_utama[kol_terpilih].dropna().astype(str).unique().tolist() 
            st.multiselect("Pilih data yang diganti (Bisa > 1):", nilai_unik, key=f"cari_teks_{uid_g}") 
            
        with col_g3: 
            st.selectbox("Ubah menjadi:", ["Teks Baru", "Kosong (NaN/Blank)"], key=f"tipe_ganti_{uid_g}") 
            
        with col_g4: 
            st.text_input("Teks Baru:", key=f"teks_baru_{uid_g}") 
            
        with col_g5: 
            st.markdown("<br>", unsafe_allow_html=True) 
            st.button("❌", key=f"hapus_ganti_{uid_g}", on_click=hapus_baris_ganti, args=(uid_g,)) 
            
    st.button("➕ Tambah Aturan Ganti Data", on_click=tambah_ganti) 

    # ------------------------------------------ 
    # UI TAHAP 2: VLOOKUP GANDA & PENEMPATAN 
    # ------------------------------------------ 
    st.markdown("---") 
    st.subheader("🔗 Tahap 2: Pengaturan VLOOKUP & Penempatan") 
    col_v1, col_v2, col_v3 = st.columns(3) 
    
    with col_v1: 
        kunci_utama = st.selectbox("Kunci di Tabel Utama:", df_utama.columns) 
    with col_v2: 
        kunci_ref = st.selectbox("Kunci di Tabel Kamus:", df_ref.columns) 
    with col_v3: 
        target_tarik = st.multiselect("Data yang Ditarik (Bisa > 1):", df_ref.columns) 

    with st.expander("🛠️ Buka Pengaturan VLOOKUP Lanjutan (Syarat & Kunci Cadangan)"): 
        st.markdown("**A. Saring Data Kamus (Contoh: Hanya ambil yang Jabatannya = CPS):**") 
        filter_kamus_aktif = st.checkbox("Aktifkan Syarat Filter Kamus") 
        if filter_kamus_aktif: 
            col_fk1, col_fk2, col_fk3 = st.columns([2, 1, 2]) 
            with col_fk1: 
                kolom_filter_kamus = st.selectbox("Kolom Syarat di Kamus:", df_ref.columns) 
            with col_fk2: 
                operator_filter_kamus = st.selectbox("Kondisi:", ["Sama Dengan (==)", "Mengandung Teks"]) 
            with col_fk3: 
                nilai_filter_kamus = st.text_input("Teks Syarat (Misal: CPS):") 
        else: 
            kolom_filter_kamus, operator_filter_kamus, nilai_filter_kamus = None, None, None 

        st.markdown("---") 
        st.markdown("**B. Kunci Cadangan / Fallback (Jika kunci pertama gagal ditemukan):**") 
        fallback_aktif = st.checkbox("Aktifkan Pencarian Kunci Cadangan") 
        if fallback_aktif: 
            col_fb1, col_fb2 = st.columns(2) 
            with col_fb1: 
                kunci_utama_cadangan = st.selectbox("Kunci Cadangan di Tabel Utama (Misal: KC):", df_utama.columns) 
            with col_fb2: 
                kunci_ref_cadangan = st.selectbox("Kunci Cadangan di Tabel Kamus (Misal: Outlet):", df_ref.columns) 
        else: 
            kunci_utama_cadangan, kunci_ref_cadangan = None, None 

    konfigurasi_tarik = {} 
    if len(target_tarik) > 0: 
        st.markdown("*Pengaturan Target Penempatan:*") 
        for col in target_tarik: 
            col_t1, col_t2 = st.columns(2) 
            with col_t1: 
                mode = st.radio(f"Penempatan untuk '{col}':", ["Buat Kolom Baru", "Timpa Kolom Lama"], key=f"mode_vlookup_{col}", horizontal=True) 
            with col_t2: 
                if mode == "Buat Kolom Baru": 
                    nama_target = st.text_input(f"Nama Kolom Baru:", value=col, key=f"nama_vlookup_{col}") 
                else: 
                    nama_target = st.selectbox(f"Pilih Kolom yang Ditimpa:", list(df_utama.columns), key=f"timpa_vlookup_{col}") 
            
            konfigurasi_tarik[col] = {'mode': mode, 'target': nama_target} 

    # ------------------------------------------ 
    # UI TAHAP 3: PERAKIT ATURAN, SPLITTER & ERASER 
    # ------------------------------------------ 
    st.subheader("⚖️ Tahap 3: Validasi & Pemisahan Data (Opsional)") 
    
    with st.expander("Buka Pengaturan Validasi & Pemisahan", expanded=True): 
        st.markdown("**A. Validasi Silang (Cek Kecocokan):**") 
        col_r1, col_r2, col_r3 = st.columns(3) 
        
        semua_target_vlookup = [conf['target'] for conf in konfigurasi_tarik.values()] if 'konfigurasi_tarik' in locals() else [] 

        kolom_vlookup_baru = [conf['target'] for conf in konfigurasi_tarik.values() if conf['mode'] == "Buat Kolom Baru"] if 'konfigurasi_tarik' in locals() else [] 
        
        kolom_gabungan_set = list(df_utama.columns) 
        for t in kolom_vlookup_baru: 
            if t not in kolom_gabungan_set: 
                kolom_gabungan_set.append(t) 
        kolom_gabungan = ["(Tidak Ada)"] + kolom_gabungan_set 
        
        with col_r1: 
            kolom_kiri = st.selectbox("Bandingkan Kolom:", kolom_gabungan) 
        with col_r2: 
            operator_logika = st.selectbox("Kondisi:", ["Sama Dengan (==)", "Tidak Sama (!=)"]) 
        with col_r3: 
            kolom_kanan = st.selectbox("Dengan Kolom:", kolom_gabungan) 
            
        col_res1, col_res2, col_res3 = st.columns(3) 
        with col_res1: 
            nama_kolom_hasil = st.text_input("Nama Kolom Hasil:", value="<<cek>>") 
        with col_res2: 
            nilai_benar = st.text_input("Jika Benar, isi:", value="0") 
        with col_res3: 
            nilai_salah = st.text_input("Jika Salah, isi:", value="1") 
            
        st.markdown("---") 
        
        st.markdown("**B. Pemisahan File (Splitter) - Kondisi Dinamis (OR):**") 
        kolom_gabungan_pisah = kolom_gabungan + [nama_kolom_hasil] 
        
        for uid in st.session_state.id_kondisi: 
            col_s1, col_s2, col_s3, col_s4 = st.columns([3, 2, 3, 1]) 
            with col_s1: 
                st.selectbox("Pisahkan JIKA kolom:", kolom_gabungan_pisah, key=f"kolom_pisah_{uid}") 
            with col_s2: 
                st.selectbox("Kondisi:", ["Mengandung Teks", "Kosong (NaN/Blank)"], key=f"tipe_pisah_{uid}") 
            with col_s3: 
                st.text_input("Nilai/Teks:", value="", key=f"nilai_pisah_{uid}") 
            with col_s4: 
                st.markdown("<br>", unsafe_allow_html=True) 
                st.button("❌", key=f"hapus_{uid}", on_click=hapus_kondisi, args=(uid,)) 
        st.button("➕ Tambah Kondisi Pisah", on_click=tambah_kondisi) 
        
        st.markdown("---") 

        st.markdown("**C. Aksi Lanjutan: Modifikasi Data (Setelah File Dipisah):**") 
        kolom_gabungan_semua = kolom_gabungan + [nama_kolom_hasil] 
        opsi_modifikasi = [c for c in kolom_gabungan_semua if c != "(Tidak Ada)"] 
        
        for uid_h in st.session_state.id_hapus: 
            col_h1, col_h2, col_h3, col_h4, col_h5 = st.columns([2.5, 2.5, 2.5, 2.5, 1]) 
            with col_h1: 
                st.multiselect("Ubah data di kolom:", opsi_modifikasi, key=f"kolom_dikosongkan_{uid_h}") 
            with col_h2: 
                st.selectbox("JIKA kolom:", kolom_gabungan_semua, key=f"kolom_syarat_kosong_{uid_h}") 
            with col_h3: 
                st.selectbox("Kondisi:", ["Sama Dengan (==)", "Mengandung Teks", "Kosong (NaN/Blank)"], key=f"kondisi_syarat_{uid_h}")
                st.text_input("Teks Syarat:", key=f"nilai_syarat_kosong_{uid_h}") 
            with col_h4: 
                st.selectbox("Ubah Menjadi:", ["Kosong (NaN/Blank)", "Teks Baru"], key=f"tipe_ubah_{uid_h}") 
                st.text_input("Teks Baru:", key=f"nilai_baru_{uid_h}") 
            with col_h5: 
                st.markdown("<br><br>", unsafe_allow_html=True) 
                st.button("❌", key=f"hapus_eraser_{uid_h}", on_click=hapus_baris_hapus, args=(uid_h,)) 
                
        st.button("➕ Tambah Aturan Modifikasi", on_click=tambah_hapus) 

    # ------------------------------------------ 
    # UI TAHAP 4: KOSMETIK, FORMATTING & RE-INDEX 
    # ------------------------------------------ 
    st.subheader("🎨 Tahap 4: Kosmetik & Penataan (Opsional)") 
    
    with st.expander("Buka Pengaturan Kosmetik"): 
        st.markdown("**Pengurutan Data (Sorting):**") 
        opsi_sorting = [col for col in kolom_gabungan_pisah if col != "(Tidak Ada)"] 
        
        col_sort1, col_sort2 = st.columns([2, 1]) 
        with col_sort1: 
            kolom_sorting = st.multiselect("Urutkan berdasarkan kolom (Urutan pilihan menentukan prioritas):", opsi_sorting) 
        with col_sort2: 
            arah_sorting = st.radio("Arah Urutan:", ["A-Z (Terkecil ke Terbesar)", "Z-A (Terbesar ke Terkecil)"]) 
        st.markdown("---") 

        st.markdown("**Tata Letak Kolom (Re-indexing):**") 
        kolom_patokan = st.selectbox( 
            "Sisipkan hasil VLOOKUP & Validasi tepat di sebelah KANAN kolom:",  
            ["(Biarkan di ujung kanan)"] + list(df_utama.columns) 
        ) 
        st.markdown("---") 

        kolom_teks_mutlak = st.multiselect("Kolom yang WAJIB TEKS (Angka nol tidak hilang):", list(df_utama.columns) + kolom_vlookup_baru)
        st.markdown("---") 
        
        st.markdown("**Pengecatan Warna Otomatis (Dinamis):**") 
        opsi_warna = [col for col in kolom_gabungan_pisah if col != "(Tidak Ada)"] 
        
        for uid_w in st.session_state.id_warna: 
            col_w1, col_w2, col_w3, col_w4 = st.columns([4, 2, 2, 1]) 
            with col_w1: 
                st.multiselect("Pilih kolom:", opsi_warna, key=f"kolom_warna_{uid_w}") 
            with col_w2: 
                st.color_picker("Warna Judul:", "#548235", key=f"warna_judul_{uid_w}") 
            with col_w3: 
                st.color_picker("Warna Isi:", "#E2EFDA", key=f"warna_isi_{uid_w}") 
            with col_w4: 
                st.markdown("<br>", unsafe_allow_html=True) 
                st.button("❌", key=f"hapus_warna_{uid_w}", on_click=hapus_baris_warna, args=(uid_w,)) 
                
        st.button("➕ Tambah Aturan Warna", on_click=tambah_warna) 

    # ========================================== 
    # TAHAP 5: MESIN EKSEKUSI (SIMPAN KE BRANKAS) 
    # ========================================== 
    st.markdown("---") 
    if st.button("🚀 JALANKAN OTOMATISASI SEKARANG!", use_container_width=True): 
        with st.spinner("Mesin sedang merakit dan memproses data..."): 
            try: 
                # 1. CLEANING AWAL & REPLACE LAMA 
                if pilihan_hapus_kosong != "(Lewati)": 
                    df_utama[pilihan_hapus_kosong] = df_utama[pilihan_hapus_kosong].fillna('').astype(str).str.strip() 
                    df_utama = df_utama[df_utama[pilihan_hapus_kosong] != ''] 

                for uid_g in st.session_state.id_ganti: 
                    kol_g = st.session_state[f"kolom_ganti_{uid_g}"] 
                    cari_list = st.session_state[f"cari_teks_{uid_g}"]  
                    tipe_g = st.session_state[f"tipe_ganti_{uid_g}"] 
                    teks_b = st.session_state[f"teks_baru_{uid_g}"] 
                    
                    if len(cari_list) > 0:  
                        df_utama[kol_g] = df_utama[kol_g].astype(str) 
                        if tipe_g == "Kosong (NaN/Blank)": 
                            df_utama[kol_g] = df_utama[kol_g].replace(cari_list, np.nan) 
                        else: 
                            df_utama[kol_g] = df_utama[kol_g].replace(cari_list, teks_b) 

                # ---> PERUBAHAN MESIN: MENAMBAHKAN .str.upper() AGAR KEY KEBAl HURUF KAPITAL <--- 
                # 2. VLOOKUP (.MAP) 
                df_utama[kunci_utama] = df_utama[kunci_utama].fillna('').astype(str).str.strip().str.upper() 
                if fallback_aktif: 
                    df_utama[kunci_utama_cadangan] = df_utama[kunci_utama_cadangan].fillna('').astype(str).str.strip().str.upper()

                df_ref[kunci_ref] = df_ref[kunci_ref].fillna('').astype(str).str.strip().str.upper() 
                if fallback_aktif: 
                    df_ref[kunci_ref_cadangan] = df_ref[kunci_ref_cadangan].fillna('').astype(str).str.strip().str.upper() 
                # ---> BATAS PERUBAHAN MESIN <--- 
                
                df_ref_bersih = df_ref.copy() 
                if filter_kamus_aktif and nilai_filter_kamus.strip() != "": 
                    kol_filter_bersih = df_ref_bersih[kolom_filter_kamus].fillna('').astype(str).str.strip().str.upper() 
                    val_filter_bersih = nilai_filter_kamus.strip().upper() 
                    
                    if operator_filter_kamus == "Sama Dengan (==)": 
                        df_ref_bersih = df_ref_bersih[kol_filter_bersih == val_filter_bersih] 
                    else: 
                        df_ref_bersih = df_ref_bersih[kol_filter_bersih.str.contains(val_filter_bersih, case=False, regex=False)] 

                df_hasil = df_utama.copy() 
                
                for col_ref, config in konfigurasi_tarik.items(): 
                    target_col = config['target'] 
                    mode = config['mode'] 
                    
                    dict_primary = df_ref_bersih.drop_duplicates(subset=[kunci_ref]).set_index(kunci_ref)[col_ref].to_dict() 
                    
                    if fallback_aktif: 
                        dict_fallback = df_ref_bersih.drop_duplicates(subset=[kunci_ref_cadangan]).set_index(kunci_ref_cadangan)[col_ref].to_dict()
                    
                    if mode == "Buat Kolom Baru": 
                        temp_series = df_hasil[kunci_utama].map(dict_primary) 
                        
                        if fallback_aktif: 
                            fallback_series = df_hasil[kunci_utama_cadangan].map(dict_fallback) 
                            temp_series = temp_series.fillna(fallback_series) 
                            
                        df_hasil[target_col] = temp_series 
                        
                    else: 
                        matched_mask_primary = df_hasil[kunci_utama].isin(dict_primary.keys()) 
                        df_hasil.loc[matched_mask_primary, target_col] = df_hasil.loc[matched_mask_primary, kunci_utama].map(dict_primary) 
                        
                        if fallback_aktif: 
                            matched_mask_fallback = (~matched_mask_primary) & df_hasil[kunci_utama_cadangan].isin(dict_fallback.keys()) 
                            df_hasil.loc[matched_mask_fallback, target_col] = df_hasil.loc[matched_mask_fallback, kunci_utama_cadangan].map(dict_fallback) 

                # 3. RULE BUILDER (Validasi) 
                if kolom_kiri != "(Tidak Ada)" and kolom_kanan != "(Tidak Ada)": 
                    kiri_bersih = df_hasil[kolom_kiri].fillna('').astype(str).str.strip().str.upper() 
                    kanan_bersih = df_hasil[kolom_kanan].fillna('').astype(str).str.strip().str.upper() 
                    
                    if operator_logika == "Sama Dengan (==)": 
                        kondisi = (kiri_bersih == kanan_bersih) 
                    else: 
                        kondisi = (kiri_bersih != kanan_bersih) 
                    
                    df_hasil[nama_kolom_hasil] = nilai_salah 
                    df_hasil.loc[kondisi, nama_kolom_hasil] = nilai_benar 

                # 4. RE-INDEXING (Geser Posisi Kolom) 
                if kolom_patokan != "(Biarkan di ujung kanan)": 
                    semua_kolom = list(df_hasil.columns) 
                    
                    kolom_vlookup_baru_eksekusi = [conf['target'] for conf in konfigurasi_tarik.values() if conf['mode'] == "Buat Kolom Baru"]
                    
                    kolom_pindahan = list(kolom_vlookup_baru_eksekusi) 
                    if kolom_kiri != "(Tidak Ada)" and nama_kolom_hasil in df_hasil.columns: 
                        kolom_pindahan.append(nama_kolom_hasil) 
                        
                    for kp in kolom_pindahan: 
                        if kp in semua_kolom: 
                            semua_kolom.remove(kp) 
                            
                    if kolom_patokan in semua_kolom: 
                        idx_patokan = semua_kolom.index(kolom_patokan) + 1 
                        for kp in kolom_pindahan: 
                            if kp in df_hasil.columns: 
                                semua_kolom.insert(idx_patokan, kp) 
                                idx_patokan += 1 
                                
                    df_hasil = df_hasil[semua_kolom] 

                # 5. PELINDUNG TIPE DATA 
                for col in df_hasil.columns: 
                    if col not in kolom_teks_mutlak: 
                        df_hasil[col] = pd.to_numeric(df_hasil[col], errors='ignore') 

                # 6. SORTING DATA 
                if len(kolom_sorting) > 0: 
                    kolom_sort_valid = [k for k in kolom_sorting if k in df_hasil.columns] 
                    if len(kolom_sort_valid) > 0: 
                        is_asc = True if "A-Z" in arah_sorting else False 
                        df_hasil.sort_values(by=kolom_sort_valid, ascending=is_asc, inplace=True) 

                # 7. SPLITTER (Membelah Data DULUAN) 
                mask_final = pd.Series(False, index=df_hasil.index) 
                pemisahan_aktif = False 
                
                for uid in st.session_state.id_kondisi: 
                    kol_p = st.session_state[f"kolom_pisah_{uid}"] 
                    tipe_p = st.session_state.get(f"tipe_pisah_{uid}", "Mengandung Teks") 
                    val_p = st.session_state[f"nilai_pisah_{uid}"] 
                    
                    if kol_p != "(Tidak Ada)": 
                        if tipe_p == "Kosong (NaN/Blank)": 
                            mask_temp = df_hasil[kol_p].isna() | (df_hasil[kol_p].astype(str).str.strip() == '') | (df_hasil[kol_p].astype(str).str.strip().str.lower() == 'nan') 
                        else: 
                            mask_temp = df_hasil[kol_p].fillna('').astype(str).str.contains(str(val_p), case=False) 
                            
                        mask_final = mask_final | mask_temp 
                        pemisahan_aktif = True 

                if pemisahan_aktif: 
                    st.session_state.df_file1 = df_hasil[~mask_final].copy() 
                    st.session_state.df_file2 = df_hasil[mask_final].copy() 
                else: 
                    st.session_state.df_file1 = df_hasil.copy() 
                    st.session_state.df_file2 = pd.DataFrame() 

                # 8. MODIFIKASI DATA / ERASER (Dieksekusi SETELAH data dibelah) 
                for df_target in [st.session_state.df_file1, st.session_state.df_file2]: 
                    if df_target is not None and not df_target.empty: 
                        
                        for uid_h in st.session_state.id_hapus: 
                            kol_dik = st.session_state[f"kolom_dikosongkan_{uid_h}"] 
                            kol_sya = st.session_state[f"kolom_syarat_kosong_{uid_h}"] 
                            tipe_kondisi = st.session_state[f"kondisi_syarat_{uid_h}"] 
                            val_sya = st.session_state[f"nilai_syarat_kosong_{uid_h}"] 
                            tipe_ubah = st.session_state[f"tipe_ubah_{uid_h}"] 
                            val_baru = st.session_state[f"nilai_baru_{uid_h}"] 
                            
                            if len(kol_dik) > 0 and kol_sya != "(Tidak Ada)": 
                                if tipe_kondisi == "Kosong (NaN/Blank)" or str(val_sya).strip() != "": 
                                    if kol_sya in df_target.columns: 
                                        sya_bersih = df_target[kol_sya].fillna('').astype(str).str.strip().str.upper() 
                                        val_bersih = str(val_sya).strip().upper() 
                                        
                                        if tipe_kondisi == "Kosong (NaN/Blank)": 
                                            baris_dihapus = df_target[kol_sya].isna() | (sya_bersih == '') | (sya_bersih == 'NAN') 
                                        elif tipe_kondisi == "Sama Dengan (==)": 
                                            baris_dihapus = (sya_bersih == val_bersih) 
                                        else: 
                                            baris_dihapus = sya_bersih.str.contains(val_bersih, case=False, regex=False) 
                                        
                                        valid_kol_dik = [k for k in kol_dik if k in df_target.columns] 
                                        if len(valid_kol_dik) > 0: 
                                            if tipe_ubah == "Kosong (NaN/Blank)": 
                                                df_target.loc[baris_dihapus, valid_kol_dik] = np.nan 
                                            else: 
                                                df_target.loc[baris_dihapus, valid_kol_dik] = val_baru 

                st.session_state.proses_selesai = True 

            except Exception as e: 
                st.error(f"❌ Terjadi kesalahan mesin: {e}") 

    # ========================================== 
    # TAHAP 6: TAMPILKAN HASIL & UNDUH (DI LUAR TOMBOL) 
    # ========================================== 
    if st.session_state.proses_selesai: 
        st.success("✅ Seluruh Pemrosesan Selesai!") 
        st.write(f"📁 **Hasil File 1 ({len(st.session_state.df_file1)} baris):**") 
        st.dataframe(st.session_state.df_file1) 
        
        if not st.session_state.df_file2.empty: 
            st.write(f"⚠️ **Hasil File 2 / Terpisah ({len(st.session_state.df_file2)} baris):**") 
            st.dataframe(st.session_state.df_file2) 
            
        st.markdown("---") 
        st.subheader("📥 Unduh Hasil Akhir") 
        col_dl1, col_dl2 = st.columns(2) 
        
        def buat_excel(df, nama_sheet): 
            buffer = io.BytesIO() 
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer: 
                df.to_excel(writer, index=False, sheet_name=nama_sheet) 
                worksheet = writer.sheets[nama_sheet] 
                
                for uid_w in st.session_state.id_warna: 
                    kolom_yg_dicat = st.session_state.get(f"kolom_warna_{uid_w}", []) 
                    
                    if len(kolom_yg_dicat) > 0: 
                        hex_judul = st.session_state[f"warna_judul_{uid_w}"].lstrip('#') 
                        hex_isi = st.session_state[f"warna_isi_{uid_w}"].lstrip('#') 
                        
                        cat_judul = PatternFill(start_color=hex_judul, end_color=hex_judul, fill_type="solid") 
                        cat_isi = PatternFill(start_color=hex_isi, end_color=hex_isi, fill_type="solid") 
                        
                        for kw in kolom_yg_dicat: 
                            if kw in df.columns: 
                                idx_warna = df.columns.get_loc(kw) + 1 
                                worksheet.cell(row=1, column=idx_warna).fill = cat_judul 
                                for row_idx in range(2, len(df) + 2): 
                                    worksheet.cell(row=row_idx, column=idx_warna).fill = cat_isi 
            return buffer.getvalue() 

        with col_dl1: 
            st.download_button( 
                label="📊 Download File 1 (Data Aman)", 
                data=buat_excel(st.session_state.df_file1, "Data Aman"), 
                file_name="Universal_Hasil_1.xlsx", 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
            ) 
            
        with col_dl2: 
            if not st.session_state.df_file2.empty: 
                st.download_button( 
                    label="⚠️ Download File 2 (Data Terpisah)", 
                    data=buat_excel(st.session_state.df_file2, "Data Terpisah"), 
                    file_name="Universal_Hasil_2.xlsx", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" 
                ) 

else: 
    st.info("👈 Silakan unggah kedua file untuk memunculkan panel pengaturan.")
