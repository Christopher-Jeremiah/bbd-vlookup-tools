import streamlit as st
import pandas as pd
import io
import numpy as np
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Ultimate Data Pipeline", layout="wide")
st.title("📊 Hub Otomatisasi VLOOKUP & Pemrosesan Data")
st.write("Jalankan pembersihan data, VLOOKUP berantai, N-Tier Fallback, dan pemecahan file Excel (Splitter) secara instan tanpa perlu merakit rumus manual.")

# ==========================================
# 0. INISIALISASI MEMORI (BRANKAS)
# ==========================================
def init_state(key, default):
    if key not in st.session_state:
        st.session_state[key] = default

init_state('id_ganti', [])
init_state('id_fb', [])
init_state('fb_conds', {}) 
init_state('id_hapus', [])
init_state('id_warna', [])
init_state('id_split2', [1]) 
init_state('id_split3', [1]) 
init_state('proses_selesai', False)
init_state('hasil_tabel', {})

def tambah_item(list_name): 
    st.session_state[list_name].append(np.random.randint(10000, 99999))
    
def hapus_item(list_name, uid): 
    if uid in st.session_state[list_name]: 
        st.session_state[list_name].remove(uid)

def tambah_fb_layer():
    new_id = np.random.randint(10000, 99999)
    st.session_state.id_fb.append(new_id)
    st.session_state.fb_conds[new_id] = [np.random.randint(10000, 99999)]

def hapus_fb_layer(uid):
    if uid in st.session_state.id_fb:
        st.session_state.id_fb.remove(uid)
    if uid in st.session_state.fb_conds:
        del st.session_state.fb_conds[uid]

def tambah_fb_cond(layer_uid):
    if layer_uid not in st.session_state.fb_conds:
        st.session_state.fb_conds[layer_uid] = []
    st.session_state.fb_conds[layer_uid].append(np.random.randint(10000, 99999))

def hapus_fb_cond(layer_uid, cond_uid):
    if layer_uid in st.session_state.fb_conds:
        if cond_uid in st.session_state.fb_conds[layer_uid]:
            st.session_state.fb_conds[layer_uid].remove(cond_uid)

# ==========================================
# 1. PINTU MASUK FILE (3 KOLOM)
# ==========================================

# MANTRA CACHE: Baca Excel/CSV sekali saja, lalu simpan di RAM agar aplikasi secepat kilat!
@st.cache_data
def baca_dan_rapihan_data(file_content, file_name, header):
    if not file_content: return pd.DataFrame()
    
    # 1. Baca File
    if file_name.endswith('.csv'): 
        df = pd.read_csv(io.BytesIO(file_content), header=header, dtype=str)
    else: 
        df = pd.read_excel(io.BytesIO(file_content), header=header, dtype=str)
        
    # 2. Rapikan Tanggal
    if df.empty: return df
    kr = []
    for col in df.columns:
        cs = str(col)
        if ' 00:00:00' in cs:
            try: kr.append(pd.to_datetime(cs).strftime('%d-%b'))
            except: kr.append(cs)
        else: kr.append(cs)
    df.columns = kr
    return df

st.markdown("### 📥 Langkah 1: Upload File Base & Referensi")
col1, col2, col3 = st.columns(3)
with col1:
    file_u = st.file_uploader("1. File Utama (Data Mentah) [Wajib]", type=["xlsx", "csv"])
    head_u = st.number_input("Mulai dari Baris Header:", 0, step=1, key="h1") if file_u else 0
with col2:
    file_k = st.file_uploader("2. File Kamus Utama (Table Array) [Wajib]", type=["xlsx", "csv"])
    head_k = st.number_input("Mulai dari Baris Header:", 0, step=1, key="h2") if file_k else 0
with col3:
    file_o = st.file_uploader("3. File Referensi Tambahan / Historis (Opsional)", type=["xlsx", "csv"])
    head_o = st.number_input("Mulai dari Baris Header:", 0, step=1, key="h3") if file_o else 0

if file_u and file_k:
    try:
        # Eksekusi fungsi cache (Mesin hanya akan loading lama di tahap ini saja)
        df_u = baca_dan_rapihan_data(file_u.getvalue(), file_u.name, head_u)
        df_k = baca_dan_rapihan_data(file_k.getvalue(), file_k.name, head_k)
        df_o = baca_dan_rapihan_data(file_o.getvalue(), file_o.name, head_o) if file_o else pd.DataFrame()
        
        kolom_dinamis = list(df_u.columns)
        dict_kamus = {"Kamus Utama": df_k}
        if not df_o.empty: dict_kamus["Referensi Tambahan"] = df_o
        pilihan_kamus = list(dict_kamus.keys())
        
    except Exception as e:
        st.error(f"Gagal membaca file. Pastikan baris judul (header) sudah benar. Pesan error: {e}")
        st.stop()
        
    st.markdown("---")

    # ==========================================
    # UI TAHAP 1: PEMBERSIHAN AWAL
    # ==========================================
    st.subheader("🧹 Tahap 1: Data Cleaning (Find & Replace)")
    
    daftar_kolom_utama = ["(Lewati)"] + kolom_dinamis
    pilihan_hapus_kosong = st.selectbox("Hapus seluruh baris jika kolom ini #N/A (Misal untuk membuang baris 'Grand Total'):", daftar_kolom_utama)
    st.markdown("---")
    
    st.markdown("**Find & Replace Data Spesifik (Dieksekusi sebelum VLOOKUP):**")
    for uid in st.session_state.id_ganti:
        c1, c2, c3, c4, c5 = st.columns([2, 4, 2, 2, 1])
        with c1: kol_g = st.selectbox("Target Kolom:", kolom_dinamis, key=f"kg_{uid}")
        with c2: 
            nilai_unik = df_u[kol_g].dropna().astype(str).unique().tolist() if kol_g in df_u.columns else []
            cari_g = st.multiselect("Find (Pilih nilai):", nilai_unik, key=f"cg_{uid}")
        with c3: tipe_g = st.selectbox("Action:", ["Replace dengan Teks", "Clear Content (#N/A)"], key=f"tg_{uid}")
        with c4: ganti_g = st.text_input("Replace With:", key=f"gg_{uid}") if tipe_g == "Replace dengan Teks" else ""
        with c5: 
            st.markdown("<br>", unsafe_allow_html=True)
            st.button("❌", key=f"dx_{uid}", on_click=hapus_item, args=('id_ganti', uid))
            
    st.button("➕ Tambah Aturan Find & Replace", key="btn_add_ganti", on_click=tambah_item, args=('id_ganti',))
    st.markdown("---")

    # ==========================================
    # UI TAHAP 1.5: VLOOKUP AWAL
    # ==========================================
    st.subheader("🕰️ Tahap 1.5: VLOOKUP Referensi Tambahan (Opsional)")
    tarik_awal_aktif = False
    
    if not df_o.empty:
        tarik_awal_aktif = st.checkbox("Aktifkan VLOOKUP ke Referensi Tambahan", value=True)
        if tarik_awal_aktif:
            c1, c2, c3 = st.columns(3)
            with c1: kunci_u_awal = st.selectbox("Lookup Value (Kunci di File Utama):", kolom_dinamis, key="ku_awal")
            with c2: kunci_o_awal = st.selectbox("Table Array (Kunci di Referensi Tambahan):", df_o.columns, key="ko_awal")
            with c3: target_o_awal = st.selectbox("Return Value (Kolom yang ditarik):", df_o.columns, key="to_awal")
            
            c4, c5 = st.columns(2)
            with c4: mode_awal = st.radio("Penempatan Hasil VLOOKUP:", ["Insert Kolom Baru", "Overwrite (Timpa) Kolom Lama"], key="ma_awal", horizontal=True)
            with c5:
                if mode_awal == "Insert Kolom Baru":
                    nama_target_awal = st.text_input("Nama Kolom Barunya:", value=target_o_awal, key="nt_awal")
                    if nama_target_awal and nama_target_awal not in kolom_dinamis:
                        kolom_dinamis.append(nama_target_awal)
                else:
                    nama_target_awal = st.selectbox("Pilih Kolom yang akan di-Overwrite:", kolom_dinamis, key="nt_awal_timpa")
                    
            st.markdown("**Splitter Tahap 1:**")
            pisah_awal_aktif = st.checkbox("Extract (Pisahkan) baris yang hasil VLOOKUP-nya 'Matched' ke file Excel terpisah", value=True)
            if pisah_awal_aktif:
                nama_t1 = st.text_input("Simpan sebagai File/Tabel bernama:", value="Tabel 1 (Match Referensi Tambahan)", key="nama_t1")
    else:
        st.info("Upload File Referensi Tambahan di Langkah 1 untuk membuka menu ini.")
    st.markdown("---")

    # ==========================================
    # UI TAHAP 2: VLOOKUP UTAMA & SPLIT
    # ==========================================
    st.subheader("🔗 Tahap 2: VLOOKUP Kamus Utama & Cross-Validation")
    t2_aktif = st.checkbox("Aktifkan VLOOKUP Kamus Utama", value=True)
    if t2_aktif:
        st.markdown("**A. Parameter VLOOKUP**")
        c1, c2 = st.columns(2)
        with c1: t2_ku = st.selectbox("Lookup Value (Kunci di File Utama):", kolom_dinamis, key="t2_ku")
        with c2: t2_kk = st.selectbox("Table Array (Kunci di Kamus Utama):", df_k.columns, key="t2_kk")
        
        st.markdown("**B. Mapping Return Value**")
        t2_tarikan = st.multiselect("Pilih kolom yang ditarik (Bisa > 1):", df_k.columns, key="t2_tarik")
        t2_map_kolom = {}
        for t in t2_tarikan:
            c_m1, c_m2 = st.columns(2)
            with c_m1:
                mode_t2 = st.radio(f"Penempatan untuk '{t}':", ["Insert Kolom Baru", "Overwrite Kolom Lama"], key=f"mode_t2_{t}", horizontal=True)
            with c_m2:
                if mode_t2 == "Insert Kolom Baru":
                    nama_t2_t = st.text_input(f"Nama Kolom Barunya:", value=t, key=f"t2_n_{t}")
                    if nama_t2_t and nama_t2_t not in kolom_dinamis: kolom_dinamis.append(nama_t2_t)
                    t2_map_kolom[t] = {"mode": mode_t2, "target": nama_t2_t}
                else:
                    nama_t2_t = st.selectbox(f"Kolom yang di-Overwrite:", kolom_dinamis, key=f"t2_s_{t}")
                    t2_map_kolom[t] = {"mode": mode_t2, "target": nama_t2_t}

        st.markdown("**C. Cross-Validation (Buat kolom cek parameter. Misal: Cek jika Nama Area di laporan sama dengan di kamus)**")
        val_aktif = st.checkbox("Aktifkan Cross-Validation")
        if val_aktif:
            c1, c2, c3 = st.columns(3)
            with c1: val_kiri = st.selectbox("Kolom Kiri:", kolom_dinamis, key="vk_kiri")
            with c2: val_op = st.selectbox("Operator:", ["Sama Persis (==)", "Berbeda (!=)"], key="vk_op")
            with c3: val_kanan = st.selectbox("Kolom Kanan:", kolom_dinamis, key="vk_kanan")
            
            c4, c5, c6 = st.columns(3)
            with c4: 
                val_hasil = st.text_input("Nama Kolom Validator:", value="<<cek>>", key="vk_hasil")
                if val_hasil not in kolom_dinamis: kolom_dinamis.append(val_hasil)
            with c5: val_b = st.text_input("Value jika TRUE:", value="0", key="vk_b")
            with c6: val_s = st.text_input("Value jika FALSE:", value="1", key="vk_s")

        st.markdown("**D. Splitter Tahap 2**")
        pisah_t2_aktif = st.checkbox("Aktifkan Ekstraksi File (Splitter)", value=True)
        if pisah_t2_aktif:
            c1, c2 = st.columns([2, 3])
            default_t2 = "Tabel 2 (Data Valid)" if not df_o.empty else "Tabel 1 (Data Valid)"
            
            with c1: nama_t2 = st.text_input("Simpan sebagai File/Tabel bernama:", value=default_t2, key="nama_t2")
            with c2: 
                aksi_t2 = st.radio("Routing Baris (Data yang memenuhi syarat di bawah akan...):", 
                    [f"Dilempar ke Tahap 3 (Sisa data yang bersih mengendap di {nama_t2})", 
                     f"Diekstrak ke {nama_t2} (Sisa data dilempar ke Tahap 3)"], 
                    key="aksi_t2")
                
            logika_t2 = st.radio("Logika Filter Gabungan:", 
                                 ["Wajib penuhi SEMUA syarat (AND)", "Cukup penuhi SALAH SATU (OR)"], horizontal=True, key="logika_t2")
            
            for uid in st.session_state.id_split2:
                c1, c2, c3, c4 = st.columns([3, 2, 3, 1])
                with c1: k_val = st.selectbox("Kolom:", kolom_dinamis, key=f"s2k_{uid}")
                with c2: o_val = st.selectbox("Operator:", ["Sama Dengan (==)", "Tidak Sama (!=)", "Mengandung Teks", "TIDAK Mengandung Teks", "Kosong (Blank / #N/A)", "Tidak Kosong (Ada isinya)"], key=f"s2o_{uid}")
                with c3: 
                    if o_val in ["Sama Dengan (==)", "Tidak Sama (!=)"]:
                        if k_val in df_u.columns:
                            ops_u = ["(Input Manual)"] + list(df_u[k_val].dropna().astype(str).unique())
                            pil_v = st.selectbox("Value:", ops_u, key=f"s2v_sel_{uid}")
                            if pil_v == "(Input Manual)":
                                st.text_input("Input Manual:", key=f"s2v_man_{uid}")
                        else:
                            st.text_input("Value:", key=f"s2v_man_{uid}")
                    elif o_val not in ["Kosong (Blank / #N/A)", "Tidak Kosong (Ada isinya)"]:
                        st.text_input("Value Teks:", key=f"s2v_{uid}")
                with c4: 
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.button("❌", key=f"s2x_{uid}", on_click=hapus_item, args=('id_split2', uid))
            st.button("➕ Tambah Syarat Filter", key="btn_add_split2", on_click=tambah_item, args=('id_split2',))
    st.markdown("---")

    # ==========================================
    # UI TAHAP 3: FALLBACK & PEMISAHAN AKHIR
    # ==========================================
    st.subheader("⚠️ Tahap 3: N-Tier Fallback VLOOKUP & Split Akhir")
    t3_aktif = st.checkbox("Aktifkan Tahap 3 (Proses baris anomali yang masih #N/A)", value=True)
    if t3_aktif:
        st.markdown("**1. Mesin N-Tier Fallback (Eksekusi VLOOKUP cadangan secara berlapis):**")
        target_fb = st.selectbox("Target Kolom yang akan diselamatkan (Diisi Fallback):", kolom_dinamis, key="tfb_target")
        
        reset_target_fb = st.checkbox("Clear Content (Hapus isi) kolom target di atas sebelum Fallback pertama berjalan agar tidak skip data lama.", value=True)

        for urutan, uid in enumerate(st.session_state.id_fb):
            if uid not in st.session_state.fb_conds: st.session_state.fb_conds[uid] = []
            
            with st.container(border=True):
                st.markdown(f"**Tier Fallback ke-{urutan+1}**")
                c1, c2, c3 = st.columns([1,1,1])
                with c1: fb_kamus = st.selectbox("Pilih Kamus:", pilihan_kamus, key=f"fbk_{uid}")
                with c2: fb_ku = st.selectbox("Lookup Value Baru:", kolom_dinamis, key=f"fbku_{uid}")
                with c3: fb_kk = st.selectbox("Table Array Baru:", dict_kamus[fb_kamus].columns, key=f"fbkk_{uid}")
                
                st.markdown("*Custom Filter untuk Kamus (Saring data kamus sebelum di-VLOOKUP):*")
                logika_fb = st.radio("Logika Filter Kamus:", ["Wajib penuhi SEMUA (AND)", "Cukup penuhi SALAH SATU (OR)"], key=f"fblog_{uid}", horizontal=True)
                
                for c_uid in st.session_state.fb_conds[uid]:
                    c4, c5, c6, c7 = st.columns([3, 2, 3, 1])
                    with c4: f_sk = st.selectbox("Filter Kolom di Kamus:", ["(Tanpa Filter)"] + list(dict_kamus[fb_kamus].columns), key=f"fbsk_{uid}_{c_uid}")
                    with c5: f_op = st.selectbox("Operator:", ["Sama Dengan (==)", "Tidak Sama (!=)", "Mengandung Teks", "TIDAK Mengandung Teks", "Kosong (Blank / #N/A)", "Tidak Kosong (Ada isinya)"], key=f"fbso_{uid}_{c_uid}")
                    with c6: f_val = st.text_input("Value:", key=f"fbsv_{uid}_{c_uid}")
                    with c7: 
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.button("❌", key=f"fbx_c_{uid}_{c_uid}", on_click=hapus_fb_cond, args=(uid, c_uid))
                        
                c_add, c_del = st.columns([2, 8])
                with c_add: st.button("➕ Tambah Custom Filter", key=f"add_c_{uid}", on_click=tambah_fb_cond, args=(uid,))
                with c_del: st.button("🗑️ Hapus Tier Ini", key=f"fbx_{uid}", on_click=hapus_fb_layer, args=(uid,))
                
        st.button("➕ Tambah Tier Fallback Baru", key="btn_add_fb_layer", on_click=tambah_fb_layer)

        st.markdown("**2. Splitter Tahap 3 (Pecah sisa data menjadi 2 output):**")
        pisah_t3_aktif = st.checkbox("Aktifkan Splitter Akhir", value=True)
        if pisah_t3_aktif:
            c1, c2 = st.columns(2)
            default_t3a = "Tabel 3A (Ekstraksi 1)" if not df_o.empty else "Tabel 2A (Ekstraksi 1)"
            default_t3b = "Tabel 3B (Sisa Akhir)" if not df_o.empty else "Tabel 2B (Sisa Akhir)"
            
            with c1: nama_t3a = st.text_input("Simpan Hasil 1 sebagai:", value=default_t3a, key="nama_t3a")
            with c2: nama_t3b = st.text_input("Simpan Sisa Akhirnya sebagai:", value=default_t3b, key="nama_t3b")
            
            aksi_t3 = st.radio("Routing Baris (Data yang memenuhi syarat di bawah akan diekstrak ke):", [nama_t3a, nama_t3b], horizontal=True, key="aksi_t3")
            logika_t3 = st.radio("Logika Filter Gabungan (Tahap 3):", ["Wajib penuhi SEMUA syarat (AND)", "Cukup penuhi SALAH SATU (OR)"], horizontal=True, key="logika_t3")
            
            for uid in st.session_state.id_split3:
                c1, c2, c3, c4 = st.columns([3, 2, 3, 1])
                with c1: k_val = st.selectbox("Kolom:", kolom_dinamis, key=f"s3k_{uid}")
                with c2: o_val = st.selectbox("Operator:", ["Sama Dengan (==)", "Tidak Sama (!=)", "Mengandung Teks", "TIDAK Mengandung Teks", "Kosong (Blank / #N/A)", "Tidak Kosong (Ada isinya)"], key=f"s3o_{uid}")
                with c3: 
                    if o_val in ["Sama Dengan (==)", "Tidak Sama (!=)"]:
                        if k_val in df_u.columns:
                            ops_u = ["(Input Manual)"] + list(df_u[k_val].dropna().astype(str).unique())
                            pil_v = st.selectbox("Value:", ops_u, key=f"s3v_sel_{uid}")
                            if pil_v == "(Input Manual)":
                                st.text_input("Input Manual:", key=f"s3v_man_{uid}")
                        else:
                            st.text_input("Value:", key=f"s3v_man_{uid}")
                    elif o_val not in ["Kosong (Blank / #N/A)", "Tidak Kosong (Ada isinya)"]:
                        st.text_input("Value Teks:", key=f"s3v_{uid}")
                with c4: 
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.button("❌", key=f"s3x_{uid}", on_click=hapus_item, args=('id_split3', uid))
            st.button("➕ Tambah Syarat Filter (Tahap 3)", key="btn_add_split3", on_click=tambah_item, args=('id_split3',))
    st.markdown("---")

    # ==========================================
    # UI TAHAP 4: MODIFIKASI SPESIFIK (ERASER)
    # ==========================================
    st.subheader("🗑️ Tahap 4: Clear Content (Eraser)")
    st.write("Mengosongkan isi sel pada kolom tertentu secara spesifik SETELAH semua proses Splitter selesai.")
    
    opsi_tabel_target = ["Semua File Output"]
    if 'nama_t1' in st.session_state: opsi_tabel_target.append(st.session_state.nama_t1)
    if 'nama_t2' in st.session_state: opsi_tabel_target.append(st.session_state.nama_t2)
    if 'nama_t3a' in st.session_state: opsi_tabel_target.append(st.session_state.nama_t3a)
    if 'nama_t3b' in st.session_state: opsi_tabel_target.append(st.session_state.nama_t3b)

    for uid in st.session_state.id_hapus:
        c1, c2, c3, c4 = st.columns([2,2,2,1])
        with c1: hapus_lokasi = st.selectbox("Eksekusi di File:", opsi_tabel_target, key=f"hl_{uid}")
        with c2: hapus_target = st.selectbox("Clear Content di Kolom:", kolom_dinamis, key=f"ht_{uid}")
        with c3: hapus_syarat = st.text_input("Kondisi IF Contains (Biarkan kosong jika hapus semua baris):", key=f"hs_{uid}")
        with c4: 
            st.markdown("<br>", unsafe_allow_html=True)
            st.button("❌", key=f"hx_{uid}", on_click=hapus_item, args=('id_hapus', uid))
    st.button("➕ Tambah Aturan Eraser", key="btn_add_eraser", on_click=tambah_item, args=('id_hapus',))
    st.markdown("---")

    # ==========================================
    # UI TAHAP 5: KOSMETIK & WARNA
    # ==========================================
    st.subheader("🎨 Tahap 5: Cell Formatting & Layout")
    with st.expander("Buka Pengaturan Formating Excel Akhir"):
        c1, c2 = st.columns(2)
        with c1: 
            sort_kol = st.multiselect("Sort A-Z Berdasarkan Kolom:", kolom_dinamis)
            sort_asc = st.radio("Order:", ["Ascending (A-Z)", "Descending (Z-A)"]) == "Ascending (A-Z)"
        with c2:
            reindex_patokan = st.selectbox("Insert/Pindahkan semua kolom baru ke sebelah KANAN kolom ini:", ["(Abaikan, taruh di ujung kanan saja)"] + list(df_u.columns))
            teks_mutlak = st.multiselect("Ubah tipe data menjadi Text mutlak (Mencegah Excel memakan angka '0'):", kolom_dinamis)

        st.markdown("**Conditional Formatting (Fill Color):**")
        for uid in st.session_state.id_warna:
            c1, c2, c3, c4 = st.columns([3,2,2,1])
            with c1: w_kol = st.multiselect("Target Kolom:", kolom_dinamis, key=f"wk_{uid}")
            with c2: w_jud = st.color_picker("Header Color:", "#548235", key=f"wj_{uid}")
            with c3: w_isi = st.color_picker("Cell Data Color:", "#E2EFDA", key=f"ws_{uid}")
            with c4: 
                st.markdown("<br>", unsafe_allow_html=True)
                st.button("❌", key=f"wx_{uid}", on_click=hapus_item, args=('id_warna', uid))
        st.button("➕ Tambah Warna", key="btn_add_warna", on_click=tambah_item, args=('id_warna',))

    # ==========================================
    # MESIN EKSEKUSI PIPELINE
    # ==========================================
    st.markdown("---")
    if st.button("🚀 EKSEKUSI PIPELINE SEKARANG!", use_container_width=True):
        with st.spinner("Mesin pintar sedang membaca parameter Anda dan mengeksekusi VLOOKUP..."):
            try:
                df_run = df_u.copy()
                st.session_state.hasil_tabel = {}

                def clean_str(s_col): 
                    return s_col.fillna('').astype(str).str.strip().str.upper()

                def build_mask(df, kol, op, val):
                    if kol not in df.columns: return pd.Series(False, index=df.index)
                    k_str = clean_str(df[kol])
                    v_str = str(val).strip().upper()
                    if "Kosong (Blank" in op: return df[kol].isna() | (k_str == '') | (k_str == 'NAN')
                    if "Tidak Kosong" in op: return df[kol].notna() & (k_str != '') & (k_str != 'NAN')
                    if "Sama Dengan" in op: return k_str == v_str
                    if "Tidak Sama" in op: return k_str != v_str
                    if "TIDAK Mengandung" in op: return ~k_str.str.contains(v_str, case=False, regex=False)
                    if "Mengandung" in op: return k_str.str.contains(v_str, case=False, regex=False)
                    return pd.Series(False, index=df.index)

                # 0. HAPUS BARIS KOSONG TOTAL
                if pilihan_hapus_kosong != "(Lewati)":
                    df_run[pilihan_hapus_kosong] = df_run[pilihan_hapus_kosong].fillna('').astype(str).str.strip()
                    df_run = df_run[df_run[pilihan_hapus_kosong] != '']

                # 1. FIND & REPLACE
                for uid in st.session_state.id_ganti:
                    kg = st.session_state[f"kg_{uid}"]
                    cg = st.session_state.get(f"cg_{uid}", [])
                    tg = st.session_state.get(f"tg_{uid}", "Teks Baru")
                    if kg in df_run.columns and len(cg) > 0:
                        df_run[kg] = df_run[kg].astype(str)
                        if tg == "Clear Content (#N/A)": df_run[kg] = df_run[kg].replace(cg, np.nan)
                        else: df_run[kg] = df_run[kg].replace(cg, st.session_state.get(f"gg_{uid}", ""))

                # 1.5. VLOOKUP AWAL & SPLIT 1
                if not df_o.empty and tarik_awal_aktif:
                    df_run[kunci_u_awal] = clean_str(df_run[kunci_u_awal])
                    df_o[kunci_o_awal] = clean_str(df_o[kunci_o_awal])
                    dict_awal = df_o.drop_duplicates(subset=[kunci_o_awal]).set_index(kunci_o_awal)[target_o_awal].to_dict()
                    
                    if mode_awal == "Insert Kolom Baru": df_run[nama_target_awal] = df_run[kunci_u_awal].map(dict_awal)
                    else: df_run[nama_target_awal] = df_run[kunci_u_awal].map(dict_awal).fillna(df_run[nama_target_awal])
                        
                    if pisah_awal_aktif and nama_target_awal in df_run.columns:
                        mask_t1 = build_mask(df_run, nama_target_awal, "Tidak Kosong", "")
                        st.session_state.hasil_tabel[nama_t1] = df_run[mask_t1].copy()
                        df_run = df_run[~mask_t1].copy()

                # 2. VLOOKUP UTAMA, VALIDASI & SPLIT 2
                if not df_run.empty and t2_aktif:
                    df_run[t2_ku] = clean_str(df_run[t2_ku])
                    df_k_t2 = df_k.copy()
                    df_k_t2[t2_kk] = clean_str(df_k_t2[t2_kk])
                    
                    for t, conf in t2_map_kolom.items():
                        d_k = df_k_t2.drop_duplicates(subset=[t2_kk]).set_index(t2_kk)[t].to_dict()
                        if conf['mode'] == "Insert Kolom Baru":
                            df_run[conf['target']] = df_run[t2_ku].map(d_k)
                        else:
                            df_run[conf['target']] = df_run[t2_ku].map(d_k).fillna(df_run[conf['target']])
                        
                    if val_aktif:
                        vk, vn = clean_str(df_run[val_kiri]), clean_str(df_run[val_kanan])
                        kondisi_v = (vk == vn) if "Sama Persis" in val_op else (vk != vn)
                        df_run[val_hasil] = val_s
                        df_run.loc[kondisi_v, val_hasil] = val_b

                    if pisah_t2_aktif:
                        if len(st.session_state.id_split2) > 0:
                            is_and = ("SEMUA" in logika_t2)
                            mask_t2 = pd.Series(True, index=df_run.index) if is_and else pd.Series(False, index=df_run.index)
                            
                            for uid in st.session_state.id_split2:
                                k = st.session_state.get(f"s2k_{uid}", "")
                                o = st.session_state.get(f"s2o_{uid}", "")
                                
                                if "Sama Dengan" in o or "Tidak Sama" in o:
                                    if k in df_u.columns:
                                        sel_v = st.session_state.get(f"s2v_sel_{uid}", "")
                                        v = st.session_state.get(f"s2v_man_{uid}", "") if sel_v == "(Input Manual)" else sel_v
                                    else:
                                        v = st.session_state.get(f"s2v_man_{uid}", "")
                                elif "Kosong" in o:
                                    v = ""
                                else:
                                    v = st.session_state.get(f"s2v_{uid}", "")
                                    
                                m_temp = build_mask(df_run, k, o, v)
                                mask_t2 = (mask_t2 & m_temp) if is_and else (mask_t2 | m_temp)
                        else:
                            mask_t2 = pd.Series(True, index=df_run.index)
                            
                        aksi_t2 = st.session_state.get("aksi_t2", "")
                        if "Dilempar ke Tahap 3" in aksi_t2:
                            st.session_state.hasil_tabel[nama_t2] = df_run[~mask_t2].copy() 
                            df_run = df_run[mask_t2].copy() 
                        else:
                            st.session_state.hasil_tabel[nama_t2] = df_run[mask_t2].copy()
                            df_run = df_run[~mask_t2].copy()

                # 3. FALLBACK & SPLIT 3
                if not df_run.empty and t3_aktif:
                    if target_fb in df_run.columns:
                        
                        if reset_target_fb:
                            df_run[target_fb] = np.nan
                        
                        for uid in st.session_state.id_fb:
                            f_kam = st.session_state[f"fbk_{uid}"]
                            f_ku = st.session_state[f"fbku_{uid}"]
                            f_kk = st.session_state[f"fbkk_{uid}"]
                            
                            df_run[f_ku] = clean_str(df_run[f_ku])
                            df_f = dict_kamus[f_kam].copy() if f_kam in dict_kamus else df_k.copy()
                            df_f[f_kk] = clean_str(df_f[f_kk])
                            
                            cond_list = st.session_state.fb_conds.get(uid, [])
                            if len(cond_list) > 0:
                                is_and_fb = ("SEMUA" in st.session_state.get(f"fblog_{uid}", ""))
                                mask_fb = pd.Series(True, index=df_f.index) if is_and_fb else pd.Series(False, index=df_f.index)
                                filter_aktif = False
                                
                                for c_uid in cond_list:
                                    f_sk = st.session_state.get(f"fbsk_{uid}_{c_uid}", "(Tanpa Filter)")
                                    if f_sk != "(Tanpa Filter)":
                                        filter_aktif = True
                                        f_op = st.session_state.get(f"fbso_{uid}_{c_uid}", "")
                                        f_val = st.session_state.get(f"fbsv_{uid}_{c_uid}", "")
                                        m_temp = build_mask(df_f, f_sk, f_op, f_val)
                                        mask_fb = (mask_fb & m_temp) if is_and_fb else (mask_fb | m_temp)
                                        
                                if filter_aktif:
                                    df_f = df_f[mask_fb]
                            
                            col_tarik_asli = t2_tarikan[0] if (t2_aktif and len(t2_tarikan)>0) else target_fb
                            try: d_dict = df_f.drop_duplicates(subset=[f_kk]).set_index(f_kk)[col_tarik_asli].to_dict()
                            except: d_dict = {}
                            
                            mk = build_mask(df_run, target_fb, "Kosong (Blank)", "")
                            df_run.loc[mk, target_fb] = df_run.loc[mk, f_ku].map(d_dict)

                    if pisah_t3_aktif:
                        if len(st.session_state.id_split3) > 0:
                            is_and = ("SEMUA" in logika_t3)
                            mask_t3 = pd.Series(True, index=df_run.index) if is_and else pd.Series(False, index=df_run.index)
                            
                            for uid in st.session_state.id_split3:
                                k = st.session_state.get(f"s3k_{uid}", "")
                                o = st.session_state.get(f"s3o_{uid}", "")
                                
                                if "Sama Dengan" in o or "Tidak Sama" in o:
                                    if k in df_u.columns:
                                        sel_v = st.session_state.get(f"s3v_sel_{uid}", "")
                                        v = st.session_state.get(f"s3v_man_{uid}", "") if sel_v == "(Input Manual)" else sel_v
                                    else:
                                        v = st.session_state.get(f"s3v_man_{uid}", "")
                                elif "Kosong" in o:
                                    v = ""
                                else:
                                    v = st.session_state.get(f"s3v_{uid}", "")
                                    
                                m_temp = build_mask(df_run, k, o, v)
                                mask_t3 = (mask_t3 & m_temp) if is_and else (mask_t3 | m_temp)
                        else:
                            mask_t3 = pd.Series(True, index=df_run.index)
                            
                        aksi_t3 = st.session_state.get("aksi_t3", "")
                        if aksi_t3 == nama_t3a:
                            st.session_state.hasil_tabel[nama_t3a] = df_run[mask_t3].copy()
                            st.session_state.hasil_tabel[nama_t3b] = df_run[~mask_t3].copy()
                        else:
                            st.session_state.hasil_tabel[nama_t3a] = df_run[~mask_t3].copy()
                            st.session_state.hasil_tabel[nama_t3b] = df_run[mask_t3].copy()
                        df_run = pd.DataFrame() 
                    else:
                        st.session_state.hasil_tabel["Data Output Akhir"] = df_run
                        df_run = pd.DataFrame()

                # 4. MODIFIKASI SPESIFIK / ERASER
                for uid in st.session_state.id_hapus:
                    h_lokasi = st.session_state[f"hl_{uid}"]
                    h_targ = st.session_state[f"ht_{uid}"]
                    h_syar = st.session_state[f"hs_{uid}"].strip().upper()
                    
                    target_dfs = []
                    if h_lokasi == "Semua File Output": target_dfs = list(st.session_state.hasil_tabel.keys())
                    elif h_lokasi in st.session_state.hasil_tabel: target_dfs = [h_lokasi]
                    
                    for t_name in target_dfs:
                        df_target = st.session_state.hasil_tabel[t_name]
                        if h_targ in df_target.columns:
                            if h_syar == "": 
                                df_target[h_targ] = np.nan
                            else: 
                                df_target.loc[clean_str(df_target[h_targ]).str.contains(h_syar, case=False, regex=False), h_targ] = np.nan
                            st.session_state.hasil_tabel[t_name] = df_target

                # 5. KOSMETIK (Re-index, Tipe Data, Sort)
                for nama_tab, df_res in st.session_state.hasil_tabel.items():
                    if not df_res.empty:
                        if "(Abaikan" not in reindex_patokan:
                            cols = list(df_u.columns)
                            new_cols = [c for c in df_res.columns if c not in cols]
                            if reindex_patokan in cols:
                                idx_p = cols.index(reindex_patokan) + 1
                                for nc in new_cols:
                                    cols.insert(idx_p, nc)
                                    idx_p += 1
                            df_res = df_res[[c for c in cols if c in df_res.columns]]

                        for c in df_res.columns:
                            if c not in teks_mutlak: df_res[c] = pd.to_numeric(df_res[c], errors='ignore')

                        if sort_kol:
                            valid_sort = [k for k in sort_kol if k in df_res.columns]
                            if valid_sort: df_res.sort_values(by=valid_sort, ascending=sort_asc, inplace=True, na_position='last')
                            
                        st.session_state.hasil_tabel[nama_tab] = df_res

                st.session_state.proses_selesai = True

            except Exception as e:
                import traceback
                st.error(f"❌ Terjadi kesalahan pada mesin. Detail error: {traceback.format_exc()}")

    # ==========================================
    # UNDUH HASIL
    # ==========================================
    if st.session_state.proses_selesai:
        # PELINDUNG 1: Jika tidak ada pemecahan tabel yang aktif
        if len(st.session_state.hasil_tabel) == 0:
            st.session_state.hasil_tabel["Data Output Akhir"] = df_run

        st.success("✅ Yey! Data Anda sudah selesai diproses. Silakan unduh file-nya di bawah ini.")
        st.markdown("---")
        
        def buat_excel(df):
            b = io.BytesIO()
            with pd.ExcelWriter(b, engine='openpyxl') as w: 
                df.to_excel(w, index=False)
                ws = w.sheets['Sheet1']
                for uid in st.session_state.id_warna:
                    k_warn = st.session_state.get(f"wk_{uid}", [])
                    if k_warn:
                        cj = PatternFill(start_color=st.session_state[f"wj_{uid}"].lstrip('#'), end_color=st.session_state[f"wj_{uid}"].lstrip('#'), fill_type="solid")
                        ci = PatternFill(start_color=st.session_state[f"ws_{uid}"].lstrip('#'), end_color=st.session_state[f"ws_{uid}"].lstrip('#'), fill_type="solid")
                        for kw in k_warn:
                            if kw in df.columns:
                                idx_w = df.columns.get_loc(kw) + 1
                                ws.cell(row=1, column=idx_w).fill = cj
                                for r in range(2, len(df) + 2): ws.cell(row=r, column=idx_w).fill = ci
            return b.getvalue()
            
        # PELINDUNG 2: Keamanan pembuatan kolom unduh
        jumlah_tabel = len(st.session_state.hasil_tabel)
        if jumlah_tabel > 0:
            cols_dl = st.columns(jumlah_tabel)
            for idx, (nama_file, data_tabel) in enumerate(st.session_state.hasil_tabel.items()):
                with cols_dl[idx]:
                    st.info(f"📁 **{nama_file}**\nFile ini berisi {len(data_tabel)} baris data.")
                    st.download_button(label=f"⬇️ Unduh Excel", data=buat_excel(data_tabel), file_name=f"{nama_file}.xlsx", key=f"dl_{idx}")
else:
    st.info("👈 Silakan upload minimal File Utama dan Kamus Utama untuk mulai mengatur proses data.")
