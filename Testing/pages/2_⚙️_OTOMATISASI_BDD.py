import streamlit as st
import pandas as pd
import numpy as np
import io

# Impor pustaka warna dari openpyxl
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Otomatisasi BBD", page_icon="⚙️", layout="wide")
st.title("⚙️ Otomatisasi BBD")
st.write("Aplikasi ini dirancang khusus untuk memproses data BBD. Anda wajib mengunggah File Utama dan Kamus Utama (List NIP). Kamus Opsional (BBD Kemarin) bersifat opsional.")

# ==========================================
# 0. INISIALISASI BRANKAS MEMORI
# ==========================================
if 'proses_selesai' not in st.session_state:
    st.session_state.proses_selesai = False
if 'tabel_1' not in st.session_state:
    st.session_state.tabel_1 = None
if 'tabel_2' not in st.session_state:
    st.session_state.tabel_2 = None
if 'tabel_3a' not in st.session_state:
    st.session_state.tabel_3a = None
if 'tabel_3b' not in st.session_state:
    st.session_state.tabel_3b = None
if 'pakai_kamus_opsional' not in st.session_state:
    st.session_state.pakai_kamus_opsional = False

# ==========================================
# 1. PINTU MASUK DATA & PENGATURAN UI DINAMIS
# ==========================================
st.sidebar.header("📂 Upload File")
file_utama = st.sidebar.file_uploader("1. File Utama (BBD Hari Ini) *[Wajib]*", type=["xlsx", "csv"])
file_kamus_utama = st.sidebar.file_uploader("2. Kamus Utama (List NIP) *[Wajib]*", type=["xlsx", "csv"])
file_kamus_opsional = st.sidebar.file_uploader("3. Kamus Opsional (BBD Kemarin) *[Opsional]*", type=["xlsx", "csv"])

st.sidebar.markdown("---")
st.sidebar.header("⚙️ Pengaturan Pembacaan")

if file_utama is not None:
    if file_utama.name.endswith('.xlsx'):
        xls_u = pd.ExcelFile(file_utama)
        sheet_u = st.sidebar.selectbox("File Utama: Pilih Sheet", xls_u.sheet_names)
    else:
        sheet_u = 0
    header_u = st.sidebar.number_input("File Utama: Baris Header (Mulai 0)", min_value=0, value=0, step=1)
else:
    sheet_u = 0
    header_u = 1

if file_kamus_utama is not None:
    if file_kamus_utama.name.endswith('.xlsx'):
        xls_ku = pd.ExcelFile(file_kamus_utama)
        sheet_k_utama = st.sidebar.selectbox("Kamus Utama: Pilih Sheet", xls_ku.sheet_names)
    else:
        sheet_k_utama = 0
    header_k_utama = st.sidebar.number_input("Kamus Utama: Baris Header (Mulai 0)", min_value=0, value=0, step=1)
else:
    sheet_k_utama = 0
    header_k_utama = 0

if file_kamus_opsional is not None:
    if file_kamus_opsional.name.endswith('.xlsx'):
        xls_ko = pd.ExcelFile(file_kamus_opsional)
        sheet_k_opsional = st.sidebar.selectbox("Kamus Opsional: Pilih Sheet", xls_ko.sheet_names)
    else:
        sheet_k_opsional = 0
    header_k_opsional = st.sidebar.number_input("Kamus Opsional: Baris Header (Mulai 0)", min_value=0, value=1, step=1)
else:
    sheet_k_opsional = 0
    header_k_opsional = 1

@st.cache_data
def baca_file(file_content, file_name, sheet, header):
    if not file_content: return pd.DataFrame()
    if file_name.endswith('.csv'):
        return pd.read_csv(io.BytesIO(file_content), header=header, dtype=str)
    else:
        return pd.read_excel(io.BytesIO(file_content), sheet_name=sheet, header=header, dtype=str)

# ==========================================
# 2. MESIN PENGOLAHAN (CASCADING ETL)
# ==========================================
if file_utama is not None and file_kamus_utama is not None:
    if st.button("🚀 Eksekusi Proses Data Sekarang!", use_container_width=True):
        with st.spinner("Mesin sedang memproses..."):
            try:
                st.session_state.pakai_kamus_opsional = (file_kamus_opsional is not None)
                
                df_u = baca_file(file_utama.getvalue(), file_utama.name, sheet_u, header_u)
                df_k_utama = baca_file(file_kamus_utama.getvalue(), file_kamus_utama.name, sheet_k_utama, header_k_utama)
                
                def perbaiki_judul_tanggal(df):
                    kolom_rapi = []
                    for col in df.columns:
                        if isinstance(col, pd.Timestamp):
                            kolom_rapi.append(col.strftime('%d-%b').upper())
                        else:
                            col_str = str(col)
                            if '00:00:00' in col_str or (len(col_str) >= 10 and col_str[4] == '-' and col_str[7] == '-'):
                                try:
                                    tgl_obj = pd.to_datetime(col_str)
                                    kolom_rapi.append(tgl_obj.strftime('%d-%b').upper())
                                except:
                                    kolom_rapi.append(col_str.upper())
                            else:
                                kolom_rapi.append(col_str)
                    df.columns = kolom_rapi
                    return df

                df_u = perbaiki_judul_tanggal(df_u)
                
                if 'TGL REAL' in df_u.columns:
                    df_u['TGL REAL'] = pd.to_datetime(df_u['TGL REAL'], errors='coerce').dt.strftime('%d-%m-%Y').fillna('')

                # =========================================================
                # Pembersihan Grand Total Menggunakan GP1PDT
                # =========================================================
                if 'GP1PDT' in df_u.columns:
                    df_u['GP1PDT'] = df_u['GP1PDT'].fillna('').astype(str).str.strip()
                    df_u = df_u[(df_u['GP1PDT'] != '') & (df_u['GP1PDT'].str.lower() != 'nan') & (~df_u['GP1PDT'].str.lower().str.contains('total'))]

                
                df_u['NIP RM'] = df_u['NIP RM'].fillna('').astype(str).str.strip()
                df_u['ACCOUNT'] = df_u['ACCOUNT'].fillna('').astype(str).str.strip()

                df_k_utama['NIP'] = df_k_utama['NIP'].fillna('').astype(str).str.strip()
                df_k_utama['K_Outlet'] = df_k_utama['K_Outlet'].fillna('').astype(str).str.strip()
                df_k_utama['Outlet'] = df_k_utama['Outlet'].fillna('').astype(str).str.strip().str.upper()
                if 'Kantor Cabang' in df_k_utama.columns:
                    df_k_utama['Kantor Cabang'] = df_k_utama['Kantor Cabang'].fillna('').astype(str).str.strip().str.upper()
                df_k_utama['Jabatan'] = df_k_utama['Jabatan'].fillna('').astype(str).str.strip().str.upper()

                # =========================================================
                # PROSES 1: MENGGUNAKAN DATA KEMARIN (JIKA ADA)
                # =========================================================
                if 'KONVERSI KPP' in df_u.columns:
                    dict_replace_kpp = {
                        "KPP Demand - Konversi Konsumer": "Konversi Konsumer",
                        "KPP Demand Konversi KPR - CP RM": "Konversi Konsumer",
                        "KPP Demand - Kompensasi Konsumer": "Konversi Konsumer"
                    }
                    df_u['KONVERSI KPP'] = df_u['KONVERSI KPP'].replace(dict_replace_kpp)
                
                if 'RMCode' not in df_u.columns:
                    df_u['RMCode'] = np.nan
                
                if st.session_state.pakai_kamus_opsional:
                    df_k_opsional = baca_file(file_kamus_opsional.getvalue(), file_kamus_opsional.name, sheet_k_opsional, header_k_opsional)
                    df_k_opsional['ACCOUNT'] = df_k_opsional['ACCOUNT'].fillna('').astype(str).str.strip()
                    kamus_1_rmcode = df_k_opsional.drop_duplicates(subset=['ACCOUNT']).set_index('ACCOUNT')['RMCode'].to_dict()
                    df_u['RMCode'] = df_u['ACCOUNT'].map(kamus_1_rmcode)
                
                # ---> PERBAIKAN STRUKTUR KOLOM AGAR <<cek>> & K_Out TIDAK TERLEMPAR KE KANAN SAAT DIGABUNG <---
                if 'SNAME' in df_u.columns:
                    cols = list(df_u.columns)
                    if 'RMCode' in cols: cols.remove('RMCode')
                    if '<<cek>>' in cols: cols.remove('<<cek>>')
                    if 'K_Out by NIP' in cols: cols.remove('K_Out by NIP')
                    
                    idx_sname = cols.index('SNAME')
                    cols.insert(idx_sname + 1, 'RMCode')
                    cols.insert(idx_sname + 2, '<<cek>>')
                    cols.insert(idx_sname + 3, 'K_Out by NIP')
                    
                    df_u['<<cek>>'] = np.nan
                    df_u['K_Out by NIP'] = np.nan
                    df_u = df_u[cols]
                
                mask_match_acc = df_u['RMCode'].notna() & (df_u['RMCode'].astype(str).str.strip() != '') & (df_u['RMCode'].astype(str).str.strip().str.lower() != 'nan')
                tabel_1 = df_u[mask_match_acc].copy()
                tabel_2 = df_u[~mask_match_acc].copy()

                # =========================================================
                # PROSES 2: VALIDASI NIP NORMAL PADA TABEL 2
                # =========================================================
                if not tabel_2.empty:
                    kamus_2_kout = df_k_utama.drop_duplicates(subset=['NIP']).set_index('NIP')['K_Outlet'].to_dict()
                    tabel_2['K_Out by NIP'] = tabel_2['NIP RM'].map(kamus_2_kout).fillna('')
                    
                    tabel_2['BRANCH'] = tabel_2['BRANCH'].fillna('').astype(str).str.strip()
                    tabel_2['<<cek>>'] = np.where(tabel_2['K_Out by NIP'] == tabel_2['BRANCH'], '0', '1')
                    
                    kamus_2_rmcode = df_k_utama.drop_duplicates(subset=['NIP']).set_index('NIP')['RMCode'].to_dict()
                    mask_cek_0 = tabel_2['<<cek>>'] == '0'
                    
                    tabel_2['RMCode'] = tabel_2['RMCode'].astype(object) # FIX PANDAS WARNING
                    tabel_2.loc[mask_cek_0, 'RMCode'] = tabel_2.loc[mask_cek_0, 'NIP RM'].map(kamus_2_rmcode)
                    
                    mask_cek_1 = tabel_2['<<cek>>'] == '1'
                    mask_sph = tabel_2['RMCode'].fillna('').astype(str).str.contains('SPH', case=False)
                    
                    tabel_3 = tabel_2[mask_cek_1 | mask_sph].copy()
                    tabel_2_clean = tabel_2[~(mask_cek_1 | mask_sph)].copy()
                else:
                    tabel_3 = pd.DataFrame()
                    tabel_2_clean = pd.DataFrame()

                # =========================================================
                # PROSES 3: VLOOKUP BERSYARAT (5-TIER FALLBACK) PADA TABEL 3
                # =========================================================
                if not tabel_3.empty:
                    df_k_utama_cps_exact = df_k_utama[df_k_utama['Jabatan'] == 'CPS']
                    
                    df_k_utama_cps_contains = df_k_utama[
                        df_k_utama['Jabatan'].str.contains('CPS', case=False, na=False) & 
                        (df_k_utama['Jabatan'] != 'CPS') & 
                        (~df_k_utama['Jabatan'].str.contains('#', na=False))
                    ]
                    
                    kamus_branch_kout_cps = df_k_utama_cps_exact.drop_duplicates(subset=['K_Outlet']).set_index('K_Outlet')['RMCode'].to_dict()
                    kamus_kc_outlet_cps = df_k_utama_cps_exact.drop_duplicates(subset=['Outlet']).set_index('Outlet')['RMCode'].to_dict()
                    kamus_kc_outlet_smecps = df_k_utama_cps_contains.drop_duplicates(subset=['Outlet']).set_index('Outlet')['RMCode'].to_dict()
                    kamus_kc_kancab_cps = df_k_utama_cps_exact.drop_duplicates(subset=['Kantor Cabang']).set_index('Kantor Cabang')['RMCode'].to_dict()
                    kamus_kc_kancab_smecps = df_k_utama_cps_contains.drop_duplicates(subset=['Kantor Cabang']).set_index('Kantor Cabang')['RMCode'].to_dict()
                    
                    tabel_3['BRANCH'] = tabel_3['BRANCH'].fillna('').astype(str).str.strip()
                    tabel_3['KC'] = tabel_3['KC'].fillna('').astype(str).str.strip().str.upper()
                    
                    def cek_sel_kosong(df):
                        return df['RMCode'].isna() | (df['RMCode'].astype(str).str.strip() == '') | (df['RMCode'].astype(str).str.strip().str.lower() == 'nan')

                    tabel_3['RMCode'] = tabel_3['BRANCH'].map(kamus_branch_kout_cps)
                    
                    mask_kosong = cek_sel_kosong(tabel_3)
                    tabel_3.loc[mask_kosong, 'RMCode'] = tabel_3.loc[mask_kosong, 'KC'].map(kamus_kc_outlet_cps)

                    mask_kosong = cek_sel_kosong(tabel_3)
                    tabel_3.loc[mask_kosong, 'RMCode'] = tabel_3.loc[mask_kosong, 'KC'].map(kamus_kc_outlet_smecps)

                    mask_kosong = cek_sel_kosong(tabel_3)
                    tabel_3.loc[mask_kosong, 'RMCode'] = tabel_3.loc[mask_kosong, 'KC'].map(kamus_kc_kancab_cps)

                    mask_kosong = cek_sel_kosong(tabel_3)
                    tabel_3.loc[mask_kosong, 'RMCode'] = tabel_3.loc[mask_kosong, 'KC'].map(kamus_kc_kancab_smecps)

                    if 'KONVERSI KPP' in tabel_3.columns:
                        kpp_str = tabel_3['KONVERSI KPP'].fillna('').astype(str).str.strip().str.lower()
                        mask_bukan_konversi = (kpp_str != 'konversi konsumer')
                    else:
                        mask_bukan_konversi = pd.Series(True, index=tabel_3.index)
                        
                    tabel_3b = tabel_3[mask_bukan_konversi].copy()
                    tabel_3a = tabel_3[~mask_bukan_konversi].copy()
                    
                    tabel_3b['RMCode'] = np.nan
                else:
                    tabel_3a = pd.DataFrame()
                    tabel_3b = pd.DataFrame()

                def urutkan_rmcode(df):
                    if not df.empty and 'RMCode' in df.columns:
                        return df.sort_values(by='RMCode', ascending=True, na_position='last')
                    return df

                tabel_1 = urutkan_rmcode(tabel_1)
                tabel_2_clean = urutkan_rmcode(tabel_2_clean)
                tabel_3a = urutkan_rmcode(tabel_3a)
                tabel_3b = urutkan_rmcode(tabel_3b)

                # =========================================================
                # PROSES TAMBAHAN: MENGEMBALIKAN KOLOM NOMINAL MENJADI ANGKA
                # =========================================================
                def kembalikan_format_angka(df):
                    if df.empty: return df
                    for col in df.columns:
                        col_str = str(col).strip().upper()
                        
                        is_tanggal = len(col_str) == 6 and col_str[2] == '-'
                        is_nominal = 'CAIR' in col_str or col_str in ['MTDREL', 'AMTREL']
                        is_account = col_str == 'ACCOUNT'
                        
                        if is_tanggal or is_nominal or is_account:
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                    return df
                
                tabel_1 = kembalikan_format_angka(tabel_1)
                tabel_2_clean = kembalikan_format_angka(tabel_2_clean)
                tabel_3a = kembalikan_format_angka(tabel_3a)
                tabel_3b = kembalikan_format_angka(tabel_3b)

                # =========================================================

                st.session_state.tabel_1 = tabel_1
                st.session_state.tabel_2 = tabel_2_clean
                st.session_state.tabel_3a = tabel_3a
                st.session_state.tabel_3b = tabel_3b
                st.session_state.proses_selesai = True
                
            except Exception as e:
                import traceback
                st.error(f"❌ Terjadi kesalahan mesin: {e}. Detail: {traceback.format_exc()}")

# ==========================================
# 3. AREA UNDUH DATA (ADAPTIF UI DENGAN KOSMETIK WARNA)
# ==========================================
if st.session_state.proses_selesai:
    st.success("✅ Seluruh tahapan Cascading ETL, Kosmetik Tanggal, dan Pengurutan A-Z berhasil dieksekusi!")
    st.markdown("---")
    
    def df_to_excel(df):
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl', date_format='DD-MM-YYYY', datetime_format='DD-MM-YYYY') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            
            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            # --- FITUR BARU 1: HELPER VLOOKUP (Menyisipkan Baris di Atas) ---
            worksheet.insert_rows(1)
            
            col_names = [str(c).strip().upper() for c in df.columns]
            if 'ACCOUNT' in col_names:
                idx_account = col_names.index('ACCOUNT') + 1
                nomor_urut = 1
                for col_num in range(idx_account, len(df.columns) + 1):
                    worksheet.cell(row=1, column=col_num).value = nomor_urut
                    nomor_urut += 1
            # -------------------------------------------------------------
            
            # --- FITUR BARU 2: PALET WARNA (KOSMETIK) ---
            warna_hijau_tua = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            warna_hijau_muda = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            warna_kuning = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")     
            warna_pink = PatternFill(start_color="FFCCFF", end_color="FFCCFF", fill_type="solid")       
            
            # PERUBAHAN: Menambahkan 'ACCOUNT' ke dalam daftar yang diwarnai hijau
            kolom_hijau = ['ACCOUNT', 'RMCODE', '<<CEK>>', 'K_OUT BY NIP']
            # -------------------------------------------------------------

            format_ribuan = '#,##0' 
            format_akun = '0'       
            
            for col_idx, col_name in enumerate(df.columns, start=1):
                col_str = str(col_name).strip().upper()
                
                # --- APPLY KOSMETIK WARNA PADA HEADER (Baris 2) ---
                sel_header = worksheet.cell(row=2, column=col_idx)
                
                is_tanggal = len(col_str) == 6 and col_str[2] == '-'
                is_cair = 'CAIR' in col_str 
                is_hijau = col_str in kolom_hijau # Berlaku untuk ACCOUNT, RMCODE, dll.
                
                if is_hijau:
                    sel_header.fill = warna_hijau_tua
                elif is_tanggal:
                    sel_header.fill = warna_kuning
                elif is_cair:
                    sel_header.fill = warna_pink
                # --------------------------------------------------

                # --- APPLY KOSMETIK FORMAT ANGKA & WARNA ISI DATA (Baris 3 ke bawah) ---
                is_numeric_col = pd.api.types.is_numeric_dtype(df[col_name])
                
                for row_idx in range(3, len(df) + 3):
                    sel_excel = worksheet.cell(row=row_idx, column=col_idx)
                    
                    if is_hijau:
                        sel_excel.fill = warna_hijau_muda
                    
                    if is_numeric_col and pd.notna(sel_excel.value):
                        if col_str == 'ACCOUNT':
                            sel_excel.number_format = format_akun
                        else:
                            sel_excel.number_format = format_ribuan
                # ----------------------------------------------------------------------
                            
        return buffer.getvalue()

    mode_unduh = st.radio("Pilih Format Output Excel:", 
        ["📂 Pisahkan menjadi file Excel terpisah (Untuk Laporan Hari Ini)", 
         "📑 Gabungkan semua ke dalam 1 File Master (Pemisah 1 Baris Kosong - Untuk Kamus Opsional Besok)"], 
        horizontal=True)

    if "Pisahkan" in mode_unduh:
        if st.session_state.pakai_kamus_opsional:
            st.subheader("📥 Unduh Hasil Pemrosesan (4 File Terpisah)")
            col1, col2 = st.columns(2)
            col3, col4 = st.columns(2)
            
            with col1:
                st.info(f"📁 **Tabel 1 (Cocok Kamus H-1)**\nBerisi {len(st.session_state.tabel_1)} baris data.")
                if not st.session_state.tabel_1.empty:
                    st.download_button("⬇️ Download Tabel 1", data=df_to_excel(st.session_state.tabel_1), file_name="1_Tabel_Match_Kemarin.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            with col2:
                st.success(f"📁 **Tabel 2 (Cocok NIP - Clean)**\nBerisi {len(st.session_state.tabel_2)} baris data.")
                if not st.session_state.tabel_2.empty:
                    st.download_button("⬇️ Download Tabel 2", data=df_to_excel(st.session_state.tabel_2), file_name="2_Tabel_Match_NIP.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
            with col3:
                st.warning(f"⚠️ **Tabel 3A (Anomali - Konversi Konsumer)**\nBerisi {len(st.session_state.tabel_3a)} baris data.")
                if not st.session_state.tabel_3a.empty:
                    st.download_button("⬇️ Download Tabel 3A", data=df_to_excel(st.session_state.tabel_3a), file_name="3A_Tabel_Anomali_Konversi_Konsumer.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
            with col4:
                st.error(f"🛑 **Tabel 3B (Anomali - Bukan Konv. Konsumer)**\nBerisi {len(st.session_state.tabel_3b)} baris data.")
                if not st.session_state.tabel_3b.empty:
                    st.download_button("⬇️ Download Tabel 3B", data=df_to_excel(st.session_state.tabel_3b), file_name="3B_Tabel_Anomali_Bukan_Konversi.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        else:
            st.subheader("📥 Unduh Hasil Pemrosesan (3 File Terpisah)")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.success(f"📁 **Data Valid (NIP & Cabang Cocok)**\nBerisi {len(st.session_state.tabel_2)} baris data.")
                if not st.session_state.tabel_2.empty:
                    st.download_button("⬇️ Download Data Valid", data=df_to_excel(st.session_state.tabel_2), file_name="1_Data_Valid_NIP.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    
            with col2:
                st.warning(f"⚠️ **Data Anomali (Konversi Konsumer)**\nBerisi {len(st.session_state.tabel_3a)} baris data.")
                if not st.session_state.tabel_3a.empty:
                    st.download_button("⬇️ Download Anomali Konversi", data=df_to_excel(st.session_state.tabel_3a), file_name="2_Anomali_Konversi_Konsumer.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    
            with col3:
                st.error(f"🛑 **Data Anomali (Bukan Konv. Konsumer)**\nBerisi {len(st.session_state.tabel_3b)} baris data.")
                if not st.session_state.tabel_3b.empty:
                    st.download_button("⬇️ Download Anomali Bukan Konv.", data=df_to_excel(st.session_state.tabel_3b), file_name="3_Anomali_Bukan_Konversi.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    else:
        # MODE GABUNGAN
        tabel_list = []
        nama_list = []
        
        if st.session_state.pakai_kamus_opsional and st.session_state.tabel_1 is not None and not st.session_state.tabel_1.empty:
            tabel_list.append(st.session_state.tabel_1)
            nama_list.append("Tabel 1")
            
        if st.session_state.tabel_2 is not None and not st.session_state.tabel_2.empty:
            tabel_list.append(st.session_state.tabel_2)
            nama_list.append("Tabel 2")
            
        if st.session_state.tabel_3a is not None and not st.session_state.tabel_3a.empty:
            tabel_list.append(st.session_state.tabel_3a)
            nama_list.append("Tabel 3A")
            
        if st.session_state.tabel_3b is not None and not st.session_state.tabel_3b.empty:
            tabel_list.append(st.session_state.tabel_3b)
            nama_list.append("Tabel 3B")
            
        list_gabungan = []
        for df in tabel_list:
            list_gabungan.append(df)
            baris_kosong = pd.DataFrame(np.nan, index=[0], columns=df.columns)
            list_gabungan.append(baris_kosong)
            
        if list_gabungan:
            list_gabungan.pop() # Menghapus baris kosong di paling ujung bawah
            df_gabungan = pd.concat(list_gabungan, ignore_index=True)
            
            teks_info = " + ".join(nama_list)
            st.info(f"📑 **File Gabungan Master**\nBerisi penggabungan dari: {teks_info}. \n*(Total {len(df_gabungan)} baris termasuk baris pemisah kosong)*")
            
            st.download_button(label=f"⬇️ Unduh 1 File Master (Siap untuk VLOOKUP H-1 Besok)", data=df_to_excel(df_gabungan), file_name=f"1_by_Rek_All_Master_Data.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

elif not file_utama or not file_kamus_utama:
    st.info("👈 Silakan unggah minimal File Utama dan Kamus Utama (List NIP) di menu sebelah kiri terlebih dahulu.")
